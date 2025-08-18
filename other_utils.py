import csv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from io import BytesIO

from models.account_roster import RosterEntryUpload
from snowflake_utils import get_table_columns

def create_excel_template(snowflake_session, table_name, worksheet_title: str): # TODO: this needs to generalize
    """
    Need to document this fully
    """
    wb = Workbook()

    ws = wb.active

    ws.title = worksheet_title
    #TODO: need a general way to exclude fields? Maybe create model?
    # headers = get_table_columns(snowflake_session=snowflake_session, table_name = table_name)
    headers = list(RosterEntryUpload.model_fields.keys())
    dummy_data = [''] * len(headers)

    data = [headers, dummy_data]

    for row in data:
        ws.append(row)

    table = Table(displayName = "Account_Roster", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    ws.add_table(table)

    for cell in ws[1]:
        if cell.value == "DEPLOYMENT_DATE":
            comment = Comment(text="The accepted format is YYYY-MM-DD.\nIf not deployed use 9999-12-31.", author='')
            cell.comment = comment

    io = BytesIO()
    wb.save(io)

    return io


def convert_excel_to_csv(uploaded_file: BytesIO):

    wb = load_workbook(uploaded_file, data_only=True, read_only=True)

    ws = wb.active

    headers = [cell.value for cell in next(ws.rows)]

    with open("Account_Roster.csv", mode = "w") as file_handle:
        csv_writer = csv.DictWriter(file_handle, fieldnames=headers)
        csv_writer.writeheader()
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            data = dict(zip(headers, values))
            valid_record = RosterEntryUpload(**data)
            print(valid_record)
            csv_writer.writerow(valid_record.model_dump())
    return file_handle

# wb = Workbook()

# ws = wb.active

# ws.title = "worksheet_title"

# test_list = [['A', 'B', 'C'], [1, 2, 3]]

# for row in test_list:
#     ws.append(row)

# table = Table(displayName = "Account_Roster", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))

# style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
# table.tableStyleInfo = style

# ws.add_table(table)

# wb.save('test.xlsx')